#!/usr/bin/env python3
"""
macro_snapshot_export.py  (v2.2 — 适配 macro_final_v11 / v11.2 权重重构版)
Excel → JSON 快照导出脚本

用法：
  python3 macro_snapshot_export.py macro_final_v11.2.xlsx
  python3 macro_snapshot_export.py macro_final_v11.2.xlsx --output-dir ./snapshots
  python3 macro_snapshot_export.py macro_final_v11.2.xlsx --context "美伊停火谈判中，市场关注油价能否回落"

输出：
  snapshot_YYYY-MM-DD.json（完整快照）
  snapshot_YYYY-MM-DD_lite.json（精简版，给AI Prompt用）

变更记录：
  v1.0 (2026-04-15) 初版，适配v10
  v2.0 (2026-04-16) 适配v11：行号全面更新+17个新指标+manual_context
  v2.1 (2026-04-17) 适配v11重排后的新行号：NODE_INDICATORS全量重写
                    CALCULATED_ROWS扩展为{40,41,43,44,52,53,61,69,74,105,113}
                    传导链路Sheet已清空行,max_row=45(不影响CHAIN_DEFS)
  v2.2 (2026-04-17) 议题1权重重构：R6/R7/R29/R30/R35节点主项/子项重新认定
                    R6留4主项(社融/贷款/M1/M2),R7加入R4+R6主项,R29加入R91/R92/R100
                    R30/R35删SOFR子项,R35加入R116联储资产
                    R116改为52周同比打分(原为绝对值)
                    CALCULATED_ROWS新增105,116都是同比计算型
"""

import json, sys, os, glob
from datetime import datetime, date
import openpyxl

NODE_INDICATORS = {
    # ===== A股 =====
    4:  [16, 17, 19, 20, 21, 22, 23],                                # CN-cb 央行态度 (未变)
    5:  [24, 25, 26, 27, 60],                                        # CN-ib 银行间传导 (未变)
    6:  [28, 29, 38, 39],                                            # CN-fin 实体传导 (重构:留4主项,7子项权重=0)
    7:  [2, 4, 5, 6, 10, 11, 12, 13, 14, 15, 75],                    # CN-eco 经济效果 (重构:加R4/R6,删R3/R7/R8/R9)
    8:  [40, 41, 42, 43, 44, 49, 50, 51],                            # CN-stk 市场反映 (未变)
    # ===== 港股 =====
    11: [2, 4, 5, 6, 10, 11, 12, 13, 14, 15, 75],                    # HK-eco 中国经济 (引用E7,展开同步新列表)
    12: [54, 75, 79, 80, 81, 106],                                   # HK-ext 海外环境 (未变)
    13: [52, 53],                                                    # HK-flow 港股资金面 (未变)
    14: [55, 56],                                                    # HK-sent 港股情绪面 (未变)
    # ===== 短债 =====
    17: [16, 17, 19, 20, 21, 22, 23],                                # SB-cb 央行态度 (引用E4)
    18: [24, 25, 26, 27, 60],                                        # SB-ib 银行间 (引用E5)
    19: [57, 58, 59],                                                # SB-yield 短端收益率 (未变)
    20: [61, 62, 63, 64, 65, 66],                                    # SB-spread 短端估值 (未变)
    # ===== 中长债 =====
    23: [16, 17, 19, 20, 21, 22, 23],                                # LB-cb 央行态度 (引用E4)
    24: [2, 4, 5, 6, 10, 11, 12, 13, 14, 15, 75],                    # LB-eco 经济预期 (引用-E7,展开同步)
    25: [67, 68],                                                    # LB-yield 长端收益率 (未变)
    26: [69, 70, 71, 72],                                            # LB-spread 长端估值 (未变)
    # ===== 美股 =====
    29: [75, 77, 85, 87, 88, 89, 90, 91, 92, 94, 95, 97, 100],       # US-eco 美国经济 (重构:加R91/R92/R100,删R98)
    30: [102, 103, 105],                                             # US-liq 美国流动性 (重构:删R104 SOFR)
    31: [106],                                                       # US-sent 美股情绪 (未变)
    32: [107],                                                       # US-val 美股估值 (未变)
    # ===== 美债 =====
    35: [102, 116],                                                  # UB-fed 美联储政策 (重构:加R116联储资产,删R104 SOFR)
    36: [108, 109, 110, 111],                                        # UB-supply 美债供需 (未变)
    37: [114, 115],                                                  # UB-yield 美债收益率 (未变)
    38: [113],                                                       # UB-spread 美债利差 (未变)
    # ===== 黄金 =====
    41: [78, 79, 80],                                                # AU-rate 实际利率 (未变)
    42: [81],                                                        # AU-usd 美元强弱 (未变)
    43: [75, 76, 100],                                               # AU-infl 通胀预期 (未变)
    44: [73, 74, 83, 84],                                            # AU-trade 黄金交易 (未变)
}

CHAIN_DEFS = [
    {"name": "A股",   "chain_row": 9,  "node_rows": [4,5,6,7,8]},
    {"name": "港股",   "chain_row": 15, "node_rows": [11,12,13,14]},
    {"name": "短债",   "chain_row": 21, "node_rows": [17,18,19,20]},
    {"name": "中长债", "chain_row": 27, "node_rows": [23,24,25,26]},
    {"name": "美股",   "chain_row": 33, "node_rows": [29,30,31,32]},
    {"name": "美债",   "chain_row": 39, "node_rows": [35,36,37,38]},
    {"name": "黄金",   "chain_row": 45, "node_rows": [41,42,43,44]},
]

CALCULATED_ROWS = {40, 41, 43, 44, 52, 53, 61, 69, 74, 105, 113, 116}

def safe_val(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, date): return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and "#" in v: return None
    if isinstance(v, float):
        if v != v: return None
    return v

def read_history(ws_macro, ws_index, code, freq, max_months=12):
    target_ws = None
    target_row = None
    if ws_macro and code:
        for row in range(2, ws_macro.max_row + 1):
            if ws_macro.cell(row, 5).value == code:
                target_ws = ws_macro
                target_row = row
                break
    if not target_row and ws_index and code:
        for row in range(2, ws_index.max_row + 1):
            if ws_index.cell(row, 5).value == code:
                target_ws = ws_index
                target_row = row
                break
    if not target_ws or not target_row:
        return [], None
    data = []
    for col in range(6, target_ws.max_column + 1):
        d = target_ws.cell(1, col).value
        v = target_ws.cell(target_row, col).value
        if v is not None and d is not None:
            dt = d.strftime("%Y-%m-%d") if isinstance(d, (datetime, date)) else str(d)
            data.append({"date": dt, "value": safe_val(v)})
    if not data:
        return [], None
    if freq == "月度":
        return data[-max_months:], None
    else:
        cutoff = "2026-01-01"
        filtered = [d for d in data if d["date"] >= cutoff]
        today = datetime.now()
        month_prefix = today.strftime("%Y-%m")
        month_data = [d for d in filtered if d["date"].startswith(month_prefix)]
        month_start = month_data[0]["value"] if month_data else None
        return filtered, month_start

def build_indicator(ws_summary, row, ws_macro, ws_index):
    name = safe_val(ws_summary.cell(row, 4).value)
    if not name:
        return None
    freq = safe_val(ws_summary.cell(row, 5).value) or "未知"
    code = safe_val(ws_summary.cell(row, 6).value)
    current = safe_val(ws_summary.cell(row, 8).value)
    current_date = safe_val(ws_summary.cell(row, 9).value)
    previous = safe_val(ws_summary.cell(row, 10).value)
    previous_date = safe_val(ws_summary.cell(row, 11).value)
    change = safe_val(ws_summary.cell(row, 12).value)
    forecast = safe_val(ws_summary.cell(row, 13).value)
    weight = safe_val(ws_summary.cell(row, 14).value)
    score = safe_val(ws_summary.cell(row, 15).value)
    signal = safe_val(ws_summary.cell(row, 17).value)
    auto_view = safe_val(ws_summary.cell(row, 18).value)

    if row in CALCULATED_ROWS:
        history, month_start = [], None
    else:
        history, month_start = read_history(ws_macro, ws_index, code, freq)

    ind = {
        "name": name, "frequency": freq, "code": code,
        "current": current, "current_date": current_date,
        "previous": previous, "previous_date": previous_date,
        "change": change, "forecast": forecast,
        "weight": weight, "score": score, "signal": signal,
        "auto_view": auto_view, "history": history,
    }
    if freq in ("日度", "周度") and month_start is not None:
        ind["month_start"] = month_start
        if current is not None:
            try: ind["mtd_change"] = round(current - month_start, 4)
            except: pass
    return ind

def build_node(ws_chain, row, ws_summary, ws_macro, ws_index):
    node_id = safe_val(ws_chain.cell(row, 2).value)
    name = safe_val(ws_chain.cell(row, 3).value)
    question = safe_val(ws_chain.cell(row, 4).value)
    score = safe_val(ws_chain.cell(row, 5).value)
    signal = safe_val(ws_chain.cell(row, 6).value)
    highlights = safe_val(ws_chain.cell(row, 7).value)
    drags = safe_val(ws_chain.cell(row, 8).value)
    view = safe_val(ws_chain.cell(row, 9).value)

    indicator_rows = NODE_INDICATORS.get(row, [])
    indicators = []
    for ir in indicator_rows:
        ind = build_indicator(ws_summary, ir, ws_macro, ws_index)
        if ind: indicators.append(ind)

    return {
        "node_id": node_id, "name": name, "question": question,
        "score": round(score, 3) if isinstance(score, float) else score,
        "signal": signal,
        "highlights": highlights or "", "drags": drags or "",
        "view": view,
        "indicator_count": len(indicators), "indicators": indicators,
    }

def build_dimension_summary(ws_summary):
    dims = {}
    current_a, current_b = None, None
    for row in range(2, ws_summary.max_row + 1):
        a = ws_summary.cell(row, 1).value
        b = ws_summary.cell(row, 2).value
        if a: current_a = a
        if b: current_b = b
        d = ws_summary.cell(row, 4).value
        p = ws_summary.cell(row, 16).value
        if d and p is not None and isinstance(p, (int, float)):
            key = f"{current_a}|{current_b}" if current_b else current_a
            if key:
                if key not in dims:
                    dims[key] = {"asset": current_a, "dimension": current_b, "weighted_score": 0, "count": 0}
                dims[key]["weighted_score"] += p
                dims[key]["count"] += 1
    return [{"asset": v["asset"], "dimension": v["dimension"],
             "weighted_score": round(v["weighted_score"], 3),
             "indicator_count": v["count"]} for v in dims.values()]

def load_previous_snapshot(output_dir, current_date_str):
    pattern = os.path.join(output_dir, "snapshot_*.json")
    files = sorted(glob.glob(pattern))
    files = [f for f in files if current_date_str not in f and "_lite" not in f]
    if not files: return None
    with open(files[-1], "r", encoding="utf-8") as f:
        return json.load(f)

def compute_diff(current_snapshot, prev_snapshot):
    if not prev_snapshot: return None
    prev_nodes = {}
    for chain in prev_snapshot.get("chains", []):
        for node in chain.get("nodes", []):
            prev_nodes[node["node_id"]] = node
    diffs = []
    for chain in current_snapshot.get("chains", []):
        for node in chain.get("nodes", []):
            nid = node["node_id"]
            prev = prev_nodes.get(nid)
            if prev and prev.get("score") is not None and node.get("score") is not None:
                try:
                    delta = round(node["score"] - prev["score"], 3)
                    direction = "改善" if delta > 0.1 else ("走弱" if delta < -0.1 else "持平")
                    diffs.append({"node_id": nid, "name": node["name"], "chain": chain["name"],
                                  "current_score": node["score"], "previous_score": prev["score"],
                                  "delta": delta, "direction": direction})
                except: pass
    return diffs

def export(excel_path, output_dir=".", manual_context=""):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws_chain = wb['传导链路']
    ws_summary = wb['汇总']

    # 前置检查:Excel是否recalc过 (openpyxl只能读已缓存值,未打开过的公式返回None)
    sample_formulas_computed = 0
    sample_formulas_none = 0
    for r in [25, 60, 75, 105]:  # 抽查4个关键行的H列
        v = ws_summary.cell(r, 8).value
        if v is None: sample_formulas_none += 1
        elif isinstance(v, (int, float)): sample_formulas_computed += 1
    if sample_formulas_none >= 2:
        print("⚠️  警告: Excel公式似乎未被重新计算!")
        print("    openpyxl 只能读已缓存的公式值。")
        print("    请在 Excel/WPS 里打开该文件保存一次,或执行:")
        print(f"    libreoffice --headless --calc --convert-to xlsx --outdir /tmp {excel_path}")
        print("    然后用 recalc 后的文件重新运行本脚本。")
        print("    (继续执行,但可能导出大量空值)\n")

    ws_macro = wb['宏观数据']
    ws_index = wb['指数走势']
    today_str = datetime.now().strftime("%Y-%m-%d")

    meta = {
        "export_date": today_str,
        "source_file": os.path.basename(excel_path),
        "excel_version": "v11", "script_version": "v2.2",
        "total_indicators": sum(1 for r in range(2, ws_summary.max_row+1) if ws_summary.cell(r,4).value),
    }
    if manual_context:
        meta["manual_context"] = manual_context

    chains = []
    for cdef in CHAIN_DEFS:
        srow = cdef["chain_row"]
        nodes = [build_node(ws_chain, nr, ws_summary, ws_macro, ws_index) for nr in cdef["node_rows"]]
        chains.append({
            "name": cdef["name"],
            "total_score": round(safe_val(ws_chain.cell(srow, 5).value) or 0, 3),
            "signal_flow": safe_val(ws_chain.cell(srow, 6).value),
            "diagnosis": safe_val(ws_chain.cell(srow, 9).value),
            "nodes": nodes,
        })

    dim_summary = build_dimension_summary(ws_summary)
    snapshot = {"meta": meta, "chains": chains, "dimension_summary": dim_summary}

    os.makedirs(output_dir, exist_ok=True)
    prev = load_previous_snapshot(output_dir, today_str)
    if prev:
        snapshot["vs_previous"] = {"previous_date": prev["meta"]["export_date"],
                                    "node_changes": compute_diff(snapshot, prev)}
    else:
        snapshot["vs_previous"] = None

    out_path = os.path.join(output_dir, f"snapshot_{today_str}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)

    total_ind = sum(sum(len(n.get("indicators",[])) for n in c.get("nodes",[])) for c in chains)
    print(f"✅ 导出完成: {out_path}")
    print(f"   链路: {len(chains)}条 | 节点: {sum(len(c['nodes']) for c in chains)}个 | 指标条目: {total_ind}个")
    print(f"   上期对比: {'有 ('+prev['meta']['export_date']+')' if prev else '无（首次导出）'}")
    return out_path

def export_lite(full_json_path, output_path=None):
    with open(full_json_path, "r", encoding="utf-8") as f:
        snap = json.load(f)
    for chain in snap["chains"]:
        for node in chain["nodes"]:
            for ind in node.get("indicators", []):
                h = ind.get("history", [])
                ind["history"] = h[-3:] if h else []
    if not output_path:
        output_path = full_json_path.replace(".json", "_lite.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    raw = json.dumps(snap, ensure_ascii=False)
    print(f"✅ 精简版: {output_path}")
    print(f"   预估token: ~{len(raw)//3:,}")
    return output_path

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 macro_snapshot_export.py <excel文件> [--output-dir <目录>] [--context <手动背景>]")
        sys.exit(1)
    excel_path = sys.argv[1]
    output_dir = "."
    manual_context = ""
    if "--output-dir" in sys.argv:
        idx = sys.argv.index("--output-dir")
        output_dir = sys.argv[idx + 1]
    if "--context" in sys.argv:
        idx = sys.argv.index("--context")
        manual_context = sys.argv[idx + 1]
    full_path = export(excel_path, output_dir, manual_context)
    export_lite(full_path)
